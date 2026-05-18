# Copyright 2026 Mantas Jonas Marcinkevičius
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the limitations under the License.

import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.widgets import RectangleSelector
import tkinter as tk
from tkinter import ttk, messagebox

def to_sci_unicode(value, decimals=2):
    """Konvertuoja skaičių į 10ⁿ formatą su Unicode laipsniais ir kableliu."""
    if abs(value) < 1e-25: return "0"
    s = ("{:.%dE}" % decimals).format(value)
    base, exp = s.split('E')
    base = base.replace('.', ',')
    exp_int = int(exp)
    superscripts = {'0':'⁰','1':'¹','2':'²','3':'³','4':'⁴','5':'⁵','6':'⁶','7':'⁷','8':'⁸','9':'⁹','-':'⁻','+':''}
    unicode_exp = "".join(superscripts.get(c, c) for c in str(exp_int))
    return f"{base}·10{unicode_exp}"

def show_arc_width(app):
    # Gauti visas prieinamas temperatūras iš projekto
    str_to_orig = {}
    all_avail_temps = []
    for k in app.vars.keys():
        s = f"{k:g}" if isinstance(k, (int, float)) else str(k)
        str_to_orig[s] = k
        all_avail_temps.append(s)
        
    # Rūšiuojame pagal skaitinę vertę
    try:
        all_avail_temps.sort(key=lambda x: float(x.replace(',', '.')))
    except:
        all_avail_temps.sort()
        
    if not all_avail_temps:
        messagebox.showwarning("Dėmesio", "Pirmiausia užkraukite duomenų failą!")
        return
        
    # Pažymėtos temperatūros pagrindiniame lange
    selected = sorted([t for t, v in app.vars.items() if v.get()], 
                      key=lambda x: float(str(x).replace(',', '.')) if isinstance(x, (int, float)) else float(str(x).replace(',', '.')))
    
    # Nustatome pradinį pasirinkimą ir rodomas temperatūras
    if selected:
        initial_temp = "Visi pažymėti"
        plot_temps = [str_to_orig[f"{t:g}" if isinstance(t, (int, float)) else str(t)] for t in selected]
    else:
        initial_temp = all_avail_temps[0]
        plot_temps = [str_to_orig[initial_temp]]
        
    # Toplevel langas
    sw = tk.Toplevel(app.root)
    sw.title("CeraMIS – Lanko analizė (stačiakampis žymėjimas)")
    w, h = app.center_window(sw, 1700, 1350)
    
    # Pagrindinis rėmelis horizontaliam išdėstymui
    main_frame = tk.Frame(sw)
    main_frame.pack(fill=tk.BOTH, expand=True)
    
    # Kairysis rėmelis grafikui
    plot_frame = tk.Frame(main_frame)
    
    # 1. Pirmiausia pakuojame dešinįjį rėmelį (Sidebar), kad jis niekada nepasislėptų ir nesusispaustų
    sidebar = tk.Frame(main_frame, width=340, bg="#F5F5F5", bd=1, relief="solid")
    sidebar.pack(side=tk.RIGHT, fill=tk.Y, padx=0, pady=0)
    sidebar.pack_propagate(False) # Užfiksuojame plotį
    
    # 2. Tada pakuojame kairįjį rėmelį grafikui, kuris užims visą likusią laisvą erdvę
    plot_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    # Antraštės sidebare
    tk.Label(sidebar, text="LANKO ANALIZĖ", font=('Arial', 12, 'bold'), bg="#F5F5F5", fg="#1A237E").pack(pady=15)
    
    desc_text = (
        "Naudojimas:\n"
        "1. Dešiniuoju pelės mygtuku apveskite stačiakampį aplink norimą lanką.\n"
        "2. Spauskite [Užfiksuoti] arba klavišą 'S', kad išsaugotumėte lanką.\n"
        "3. Apveskite kitą lanką ir kartokite procesą."
    )
    tk.Label(sidebar, text=desc_text, font=('Arial', 9), justify="left", bg="#F5F5F5", fg="#333", wraplength=310).pack(pady=5, padx=10)
    
    # Temperatūros parinkimas (Dropdown)
    tk.Label(sidebar, text="Temperatūros parinkimas:", font=('Arial', 10, 'bold'), bg="#F5F5F5").pack(anchor="w", padx=15, pady=(10, 2))
    
    combo_values = ["Visi pažymėti"] + all_avail_temps
    
    temp_var = tk.StringVar(value=initial_temp)
    temp_combo = ttk.Combobox(sidebar, textvariable=temp_var, values=combo_values, state="readonly", font=('Arial', 10))
    temp_combo.pack(fill="x", padx=15, pady=2)
    
    # Mygtukų konteineris
    btn_frame = tk.Frame(sidebar, bg="#F5F5F5")
    btn_frame.pack(pady=15, padx=10, fill="x")
    
    # Užfiksuotų rezultatų Text laukas
    tk.Label(sidebar, text="Užfiksuoti rezultatai:", font=('Arial', 10, 'bold'), bg="#F5F5F5").pack(anchor="w", padx=15, pady=(10, 5))
    
    # Scrollable text widget
    txt_frame = tk.Frame(sidebar)
    txt_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
    txt_scroll = tk.Scrollbar(txt_frame)
    txt_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    
    res_text = tk.Text(txt_frame, height=20, width=35, font=('Consolas', 9), yscrollcommand=txt_scroll.set)
    res_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    txt_scroll.config(command=res_text.yview)
    
    res_text.insert(tk.END, "Kol kas nėra užfiksuotų lankų.\n")
    res_text.config(state=tk.DISABLED)
    
    # Duomenų rinkiniai ir Matplotlib ašis
    dpi = 100
    fig = Figure(figsize=((w - 340) / dpi, h / dpi), dpi=dpi, facecolor='white')
    ax = fig.add_subplot(111)
    
    all_datasets = {}
    cmap = plt.cm.viridis
    
    app.selector_markers = []
    app.locked_arc_artists = []
    app.locked_arc_data = []
    app.current_arc_result = None

    # 1. Pirmiausia aprašome on_select
    def on_select(eclick, erelease):
        # Valome tik laikinai nubrėžtus markerius
        for m in app.selector_markers:
            try: m.remove()
            except: pass
        app.selector_markers.clear()
        app.current_arc_result = None
        
        x1, y1 = eclick.xdata, eclick.ydata
        x2, y2 = erelease.xdata, erelease.ydata
        xmin, xmax = min(x1, x2), max(x1, x2)
        ymin, ymax = min(y1, y2), max(y1, y2)
        
        results_text = []
        temp_current_fit = []
        
        for temp, (f_data, z_n, color) in all_datasets.items():
            mask = (z_n.real >= xmin) & (z_n.real <= xmax) & \
                   (-z_n.imag >= ymin) & (-z_n.imag <= ymax)
            
            if np.any(mask):
                sel_f = f_data[mask]
                sel_zr = z_n.real[mask]
                sel_zi = -z_n.imag[mask]
                
                m = ax.plot(sel_zr, sel_zi, 'o', color=color, mec='red', mew=1.5, ms=6, alpha=0.8)[0]
                app.selector_markers.append(m)
                
                r_simple = np.max(sel_zr) - np.min(sel_zr)
                p_idx = np.argmax(sel_zi)
                fp = sel_f[p_idx]
                
                # Aproksimuojame lanką apskritimu (Least Squares)
                if len(sel_zr) >= 3:
                    A = np.c_[2 * sel_zr, 2 * sel_zi, np.ones_like(sel_zr)]
                    B = sel_zr**2 + sel_zi**2
                    C_sol, _, _, _ = np.linalg.lstsq(A, B, rcond=None)
                    xc, yc = C_sol[0], C_sol[1]
                    R_circ2 = C_sol[2] + xc**2 + yc**2
                    
                    if R_circ2 > yc**2:
                        dx = np.sqrt(R_circ2 - yc**2)
                        x_int1, x_int2 = xc - dx, xc + dx
                        r_fit = x_int2 - x_int1
                        phi = np.arcsin(min(1.0, abs(yc) / np.sqrt(R_circ2)))
                        n_cpe = 1.0 - (2.0 * phi / np.pi)
                        
                        # Nupiešiame laikiną lanką
                        x_arc = np.linspace(x_int1, x_int2, 100)
                        y_arc = yc + np.sqrt(np.clip(R_circ2 - (x_arc - xc)**2, 0, None))
                        arc_line = ax.plot(x_arc, y_arc, '--', color=color, lw=2)[0]
                        int_pts = ax.plot([x_int1, x_int2], [0, 0], 'x', color='black', ms=8, mew=2, zorder=5)[0]
                        app.selector_markers.extend([arc_line, int_pts])
                        
                        if fp > 0 and r_fit > 0:
                            omega_p = 2 * np.pi * fp
                            q_cpe = 1 / (r_fit * (omega_p ** n_cpe))
                            c_eq = 1 / (omega_p * r_fit)
                        else:
                            q_cpe, c_eq = 0, 0
                            
                        res_str = (f"{temp}K: R={to_sci_unicode(r_fit)} Ω·m, n={n_cpe:.3f}, Q={to_sci_unicode(q_cpe)}\n"
                                   f"   C_eq={to_sci_unicode(c_eq)} F/m, Int:[{to_sci_unicode(x_int1)}, {to_sci_unicode(x_int2)}]")
                        results_text.append(res_str)
                        
                        temp_current_fit.append({
                            'temp': temp, 'r_fit': r_fit, 'n_cpe': n_cpe, 'q_cpe': q_cpe, 'c_eq': c_eq,
                            'x_int1': x_int1, 'x_int2': x_int2, 'color': color,
                            'fit_x': x_arc, 'fit_y': y_arc
                        })
                    else:
                        c_simple = 1 / (2 * np.pi * fp * r_simple) if fp > 0 and r_simple > 0 else 0
                        results_text.append(f"{temp}K: Max-Min R={to_sci_unicode(r_simple)} Ω·m, C={to_sci_unicode(c_simple)} F/m")
                else:
                    c_simple = 1 / (2 * np.pi * fp * r_simple) if fp > 0 and r_simple > 0 else 0
                    results_text.append(f"{temp}K: Max-Min R={to_sci_unicode(r_simple)} Ω·m, C={to_sci_unicode(c_simple)} F/m")
        
        if not results_text:
            ax.set_title("Pasirinktoje srityje nėra taškų!")
        else:
            display_limit = 6
            display_text = "Rezultatai:\n" + "\n".join(results_text[:display_limit])
            if len(results_text) > display_limit:
                display_text += f"\n... ir dar {len(results_text)-display_limit}"
            ax.set_title(display_text, fontsize=10, fontweight='bold')
            
            # Saugome dabartinio žymėjimo duomenis, kad galėtume juos užfiksuoti
            app.current_arc_result = {
                'temp_fits': temp_current_fit,
                # Nukopijuojame gautus markerius, kad perneštume į locked
                'markers': list(app.selector_markers)
            }
        
        fig.canvas.draw()

    # 2. Antra, aprašome redraw_plot (kuris dabar taip pat iš naujo inicializuoja RectangleSelector)
    def redraw_plot(plot_temps):
        ax.clear()
        ax.set_xlabel("Z', Ω·m")
        ax.set_ylabel("-Z'', Ω·m")
        ax.grid(True, linestyle='--', alpha=0.6)
        
        ax.xaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
        ax.yaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
        ax.ticklabel_format(style='sci', scilimits=(-3, 4), axis='both')
        ax.set_aspect('auto')
        
        all_datasets.clear()
        n_sel = len(plot_temps)
        k_LA, k_AL = app._get_geometric_factors()
        
        for i, orig_temp in enumerate(plot_temps):
            f, z = app.get_filtered_data(orig_temp)
            if len(z) < 2: continue
            
            z_n = z * k_AL
            color = cmap(i / max(n_sel - 1, 1)) if n_sel > 1 else '#1f77b4'
            lbl = f"{orig_temp:g}K" if isinstance(orig_temp, (int, float)) else f"{orig_temp}K"
            ax.plot(z_n.real, -z_n.imag, 'o-', markersize=6, color=color, alpha=0.7, label=lbl)
            all_datasets[orig_temp] = (f, z_n, color)

        if len(all_datasets) <= 15:
            ax.legend(fontsize=8)
            
        # Iš naujo nupiešiame visus užfiksuotus lankus, kurie priklauso rodomoms temperatūroms
        app.locked_arc_artists.clear()
        for idx, item in enumerate(app.locked_arc_data):
            for fit in item['temp_fits']:
                matching_orig = None
                for plotted_temp in all_datasets.keys():
                    s1 = f"{plotted_temp:g}" if isinstance(plotted_temp, (int, float)) else str(plotted_temp)
                    s2 = f"{fit['temp']:g}" if isinstance(fit['temp'], (int, float)) else str(fit['temp'])
                    if s1 == s2:
                        matching_orig = plotted_temp
                        break
                        
                if matching_orig is not None:
                    color = all_datasets[matching_orig][2]
                    arc_line = ax.plot(fit['fit_x'], fit['fit_y'], '-', color=color, lw=2.5)[0]
                    int_pts = ax.plot([fit['x_int1'], fit['x_int2']], [0, 0], 'x', color='black', ms=8, mew=2, zorder=5)[0]
                    app.locked_arc_artists.extend([arc_line, int_pts])
                    
        ax.set_title(f"Dešiniuoju klavišu apveskite lanką (stačiakampis)\nUžfiksuota lankų: {len(app.locked_arc_data)}")
        
        # Privalome iš naujo atkurti RectangleSelector po ax.clear()!
        app.rs = RectangleSelector(ax, on_select,
                                     useblit=False,
                                     button=[1, 3], 
                                     minspanx=0, minspany=0,
                                     interactive=True,
                                     props=dict(facecolor='red', edgecolor='red', alpha=0.15, fill=True))
        fig.canvas.draw()

    # Užkrauname pradinį vaizdą ir sugeneruojame RectangleSelector
    redraw_plot(plot_temps)

    def on_temp_change(event):
        selected_val = temp_var.get()
        if selected_val == "Visi pažymėti":
            active_selected = [t for t, v in app.vars.items() if v.get()]
            plot_temps = [str_to_orig[f"{t:g}" if isinstance(t, (int, float)) else str(t)] for t in active_selected]
        else:
            plot_temps = [str_to_orig[selected_val]]
            
        # Valome laikiną žymėjimą prieš keičiant temperatūrą
        for m in app.selector_markers:
            try: m.remove()
            except: pass
        app.selector_markers.clear()
        app.current_arc_result = None
        
        redraw_plot(plot_temps)

    temp_combo.bind("<<ComboboxSelected>>", on_temp_change)

    def update_sidebar_text():
        res_text.config(state=tk.NORMAL)
        res_text.delete("1.0", tk.END)
        if not app.locked_arc_data:
            res_text.insert(tk.END, "Kol kas nėra užfiksuotų lankų.\n")
        else:
            for idx, item in enumerate(app.locked_arc_data):
                for fit in item['temp_fits']:
                    lbl = f"{fit['temp']:g}" if isinstance(fit['temp'], (int, float)) else str(fit['temp'])
                    res_text.insert(tk.END, f"Lankas #{idx+1} ({lbl} K):\n")
                    res_text.insert(tk.END, f"  R = {to_sci_unicode(fit['r_fit'])} Om*m\n")
                    res_text.insert(tk.END, f"  n = {fit['n_cpe']:.3f}\n")
                    res_text.insert(tk.END, f"  Q = {to_sci_unicode(fit['q_cpe'])}\n")
                    res_text.insert(tk.END, f"  C_eq = {to_sci_unicode(fit['c_eq'])} F/m\n\n")
        res_text.config(state=tk.DISABLED)
        res_text.see(tk.END)

    def lock_current_arc():
        if not app.current_arc_result:
            messagebox.showwarning("Dėmesio", "Pirmiausia dešiniuoju klavišu apveskite lanką grafike!")
            return
        
        # Pernešame laikinai nupieštus elementus į permanentinius užfiksuotus
        # Pakeičiame jų stilių į ištisinę liniją (solid)
        for m in app.current_arc_result['markers']:
            if hasattr(m, 'get_linestyle') and m.get_linestyle() == '--':
                m.set_linestyle('-')
                m.set_linewidth(2.5) # Storesnė, kad išsiskirtų
            app.locked_arc_artists.append(m)
        
        # Pašaliname iš selector_markers sąrašo, kad kitas RectangleSelector jų neištrintų
        app.selector_markers.clear()
        
        # Pridedame duomenis į užfiksuotų duomenų bazę
        app.locked_arc_data.append(app.current_arc_result)
        app.current_arc_result = None
        
        update_sidebar_text()
        ax.set_title(f"Lankas sėkmingai užfiksuotas! Apveskite kitą.\nUžfiksuota lankų: {len(app.locked_arc_data)}")
        fig.canvas.draw()

    def clear_locked_arcs():
        # Pašaliname visus nupieštus locked lanko markerius iš matplotlib ašies
        for m in app.locked_arc_artists:
            try: m.remove()
            except: pass
        app.locked_arc_artists.clear()
        
        # Išvalome ir selector markerius
        for m in app.selector_markers:
            try: m.remove()
            except: pass
        app.selector_markers.clear()
        
        app.locked_arc_data.clear()
        app.current_arc_result = None
        
        update_sidebar_text()
        ax.set_title(f"Visi užfiksuoti lankai išvalyti.\nUžfiksuota lankų: {len(app.locked_arc_data)}")
        fig.canvas.draw()

    def export_locked_arcs_to_excel():
        if not app.locked_arc_data:
            messagebox.showwarning("Dėmesio", "Nėra jokių užfiksuotų lankų eksportui!")
            return
        
        from tkinter import filedialog
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile="CeraMIS_Lanku_Rezultatai.xlsx",
            title="Išsaugoti lankų analizės rezultatus"
        )
        if not file_path: return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Lankų rezultatai"
            
            headers = [
                "Lanko indeksas", 
                "Temperatūra", 
                "Varža R", 
                "CPE rodiklis n", 
                "CPE talpa Q", 
                "Ekvivalentinė talpa C_eq", 
                "Sankirta X1", 
                "Sankirta X2"
            ]
            units = [
                "", 
                "K", 
                "Ω·m", 
                "", 
                "F·s^(n-1)/m", 
                "F/m", 
                "Ω·m", 
                "Ω·m"
            ]
            
            # Stiliai
            hdr_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
            unit_font = Font(name='Arial', size=10, italic=True, color="555555")
            hdr_fill = PatternFill("solid", fgColor="1F497D") # Graži tamsiai mėlyna
            unit_fill = PatternFill("solid", fgColor="E9EDF4") # Švelni melsva/pilkšva
            
            center_align = Alignment(horizontal="center", vertical="center")
            thin_side = Side(style="thin", color="CCCCCC")
            cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            # Įrašome antraštes (Row 1 - Origin Long Name)
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = center_align
                cell.border = cell_border
                
            # Įrašome vienetus (Row 2 - Origin Units)
            for col_idx, u in enumerate(units, 1):
                cell = ws.cell(row=2, column=col_idx, value=u)
                cell.font = unit_font
                cell.fill = unit_fill
                cell.alignment = center_align
                cell.border = cell_border
                
            # Įrašome duomenis (nuo Row 3)
            row_idx = 3
            for arc_idx, item in enumerate(app.locked_arc_data):
                for fit in item['temp_fits']:
                    data = [
                        f"Lankas #{arc_idx+1}",
                        fit['temp'],
                        fit['r_fit'],
                        fit['n_cpe'],
                        fit['q_cpe'],
                        fit['c_eq'],
                        fit['x_int1'],
                        fit['x_int2']
                    ]
                    for col_idx, val in enumerate(data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.alignment = center_align
                        cell.border = cell_border
                        
                        # Skaitinių verčių formatavimas Exceliui ir teisingas tipavimas
                        if col_idx >= 2:
                            try:
                                cell.value = float(val)
                                if col_idx in [3, 5, 6, 7, 8]:
                                    cell.number_format = '0.00E+00'
                                elif col_idx == 4:
                                    cell.number_format = '0.000'
                                else:
                                    cell.number_format = '0.0'
                            except:
                                cell.value = val
                        else:
                            cell.value = val
                    row_idx += 1
                    
            # Stulpelių pločių nustatymas
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            wb.save(file_path)
            messagebox.showinfo("Sėkmė", f"Origin-suderinami duomenys sėkmingai eksportuoti į:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Klaida", f"Nepavyko eksportuoti į Excel:\n{e}")

    # Klavišų paspaudimo įvykiai
    def on_key_press(event):
        if event.key == 's' or event.key == 'S':
            lock_current_arc()
        elif event.key == 'c' or event.key == 'C':
            clear_locked_arcs()

    fig.canvas.mpl_connect('key_press_event', on_key_press)
    
    # Priverčiame gauti fokusą klaviatūrai
    canvas_widget = FigureCanvasTkAgg(fig, master=plot_frame)
    canvas_widget.get_tk_widget().bind("<Enter>", lambda e: canvas_widget.get_tk_widget().focus_set())
    
    # Sidebar mygtukai su nuostabiu stiliumi
    btn_lock = tk.Button(btn_frame, text="🔒 Užfiksuoti lanką (S)", command=lock_current_arc,
                         bg="#2E7D32", fg="white", font=('Arial', 10, 'bold'), height=2, relief="raised", bd=3)
    btn_lock.pack(fill="x", pady=5)
    
    btn_clear = tk.Button(btn_frame, text="🧹 Valyti užfiksuotus (C)", command=clear_locked_arcs,
                          bg="#C62828", fg="white", font=('Arial', 10, 'bold'), height=2, relief="raised", bd=3)
    btn_clear.pack(fill="x", pady=5)
    
    btn_excel = tk.Button(sidebar, text="📊 Eksportuoti į Excel", command=export_locked_arcs_to_excel,
                          bg="#1565C0", fg="white", font=('Arial', 10, 'bold'), height=2, relief="raised", bd=3)
    btn_excel.pack(side=tk.BOTTOM, fill="x", padx=15, pady=15)

    # Toolbaras ir Canvas kairėje
    tb_frame = tk.Frame(plot_frame)
    tb_frame.pack(side=tk.BOTTOM, fill=tk.X)
    
    canvas = FigureCanvasTkAgg(fig, master=plot_frame)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    toolbar = NavigationToolbar2Tk(canvas, tb_frame)
    toolbar.update()
    
    fig.subplots_adjust(left=0.07, bottom=0.073, right=0.954, top=0.905, wspace=0.197, hspace=0.183)
    sw.update()
    canvas.draw()
    
    # Vieno pikselio "refresh" triukas
    def force_refresh():
        try:
            geom = sw.geometry()
            parts = geom.split('+')
            wh = parts[0].split('x')
            w_new = int(wh[0]) + 1
            h_new = int(wh[1])
            sw.geometry(f"{w_new}x{h_new}+{parts[1]}+{parts[2]}")
        except: pass
        
    sw.after(200, force_refresh)
